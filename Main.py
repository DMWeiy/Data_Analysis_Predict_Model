from tkinter import *
from tkinter import Button , messagebox
from tkinter.ttk import Combobox,Treeview, Scrollbar
import Data
import seaborn as sb
import openpyxl
import pandas as pd
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error, r2_score
import matplotlib.pyplot as plt
import numpy as np
# Run the data analysis script
Data

csv = pd.read_csv("D:\\Pycharm\\IML_Project\\New_Water_Quality_AfterCleaning.csv")
excelwriter = pd.ExcelWriter("new.xlsx")

csv.to_excel(excelwriter,index=False)
excelwriter._save()
path = "D:\\Pycharm\\IML_Project\\new.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
sheet.delete_cols(1)
workbook.save("D:\\Pycharm\\IML_Project\\new.xlsx")


# Function to load data from an Excel file into the Treeview
def load_data():
    try:
        # File path to your Excel file
        file_path = "D:\\Pycharm\\IML_Project\\new.xlsx"

        # Read the Excel file using pandas
        df = pd.read_excel(file_path)

        # Check if required columns exist
        required_columns = ['pH', 'Temperature', 'Ammonia', 'ORP', 'Water_Level','BOD', 'COD', 'DO']
        if not all(col in df.columns for col in required_columns):
            messagebox.showerror("Error", "The file does not contain the required columns.")
            return

        # Clear any existing data in the Treeview
        for item in table.get_children():
            table.delete(item)

        # Populate the Treeview with data from the DataFrame
        for index, row in df.iterrows():
            table.insert(
                "",
                END,
                values=(
                    row['pH'],
                    row['Temperature'],
                    row['Ammonia'],
                    row['ORP'],
                    row['Water_Level'],
                    row['BOD'],
                    row['COD'],
                    row['DO']
                )
            )

        # Optional: Adjust the width of the columns based on the data
        for col in ['pH', 'Temperature', 'Ammonia', 'ORP', 'Water Level','BOD', 'COD', 'DO']:
            table.column(col, width=120, anchor="center")

    except FileNotFoundError:
        messagebox.showerror("Error", "Excel file not found.")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load file. Error: {str(e)}")



def update_combobox_state():
    if plot_option.get() == 2:  # If "Box Plot" is selected
        combo_X.config(state="enabled")
        combo_Y.config(state="disabled")
    elif plot_option.get() == 1:
        combo_X.config(state="enabled")
        combo_Y.config(state="disabled")
    elif plot_option.get() == 3:
        combo_X.config(state="disabled")
        combo_Y.config(state="disabled")
    elif plot_option.get() == 4:
        combo_X.config(state="disabled")
        combo_Y.config(state="enabled")


def select_row(event):
    global selected_row_id
    selected_item = table.selection()
    if selected_item:
        # Get the selected row values
        selected_row = table.item(selected_item)["values"]

        # Populate the entry fields with the selected row values
        Temp_Input.delete(0, END)
        Temp_Input.insert(0, selected_row[1])  # Temperature
        pH_Input.delete(0, END)
        pH_Input.insert(0, selected_row[0])  # pH Value
        Am_Input.delete(0, END)
        Am_Input.insert(0, selected_row[2])  # Ammonia
        Orp_Input.delete(0, END)
        Orp_Input.insert(0, selected_row[3])  # ORP
        Wl_Input.delete(0, END)
        Wl_Input.insert(0, selected_row[4])  # Water Level
        BOD_Input.delete(0, END)
        BOD_Input.insert(0, selected_row[5])  # BOD
        COD_Input.delete(0, END)
        COD_Input.insert(0, selected_row[6])  # COB
        DO_Input.delete(0, END)
        DO_Input.insert(0, selected_row[7])  # DO

        # Store the selected item for editing

        selected_row_id = selected_item[0]  # ID of the selected row


# Function to edit the selected row with new values
def edit_button():
    if selected_row_id:
        # Get the updated values from the input fields
        updated_temp = Temp_Input.get()
        updated_ph = pH_Input.get()
        updated_am = Am_Input.get()
        updated_orp = Orp_Input.get()
        updated_wl = Wl_Input.get()
        updated_BoD = BOD_Input.get()
        updated_COB = COD_Input.get()
        updated_DO = DO_Input.get()


        # Validate input
        if not updated_temp or not updated_ph or not updated_am or not updated_orp or not updated_wl:
            messagebox.showerror("Error", "All fields must be filled!")
            return

        # Update the selected row in the Treeview
        table.item(selected_row_id, values=(updated_ph, updated_temp, updated_am, updated_orp, updated_wl,updated_BoD,updated_COB,updated_DO))

        path = "D:\\Pycharm\\IML_Project\\new.xlsx"
        try:
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active

            # Determine the correct row to update
            tree_index = table.index(selected_row_id)
            excel_row = tree_index + 2  # Assuming the first data row starts at row 2 in Excel

            # Update the corresponding Excel row
            sheet.cell(row=excel_row, column=1, value=round(float(updated_ph),2))  # Update pH value
            sheet.cell(row=excel_row, column=2, value=round(float(updated_temp),2))  # Update Temperature
            sheet.cell(row=excel_row, column=3, value=round(float(updated_am),2))  # Update Ammonia
            sheet.cell(row=excel_row, column=4, value=round(float(updated_orp),2))  # Update ORP
            sheet.cell(row=excel_row, column=5, value=round(float(updated_wl),2))  # Update Water Level
            sheet.cell(row=excel_row, column=6, value=round(float(updated_BoD), 2))
            sheet.cell(row=excel_row, column=7, value=round(float(updated_COB), 2))
            sheet.cell(row=excel_row, column=8, value=round(float(updated_DO), 2))

            workbook.save(path)
            messagebox.showinfo("Success", "Row updated successfully!")
        except FileNotFoundError:
            messagebox.showerror("Error", f"File not found at {path}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while updating the file: {str(e)}")
    else:
        messagebox.showerror("Error", "No row selected to edit.")


def delete_row():
    selected_item = table.selection()  # Get the selected item in the Treeview

    if selected_item:
        # Get the item ID of the selected row
        selected_row = table.item(selected_item)["values"]  # Extract row data from the selected item

        # Convert selected item to row index (this assumes the row index corresponds to the data)
        row_index = table.index(selected_item) + 2

        # Delete from openpyxl file
        path = "D:\\Pycharm\\IML_Project\\new.xlsx"

        # Load workbook and sheet
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        # Delete the row from Excel
        sheet.delete_rows(row_index)

        # Save the changes
        workbook.save(path)
        # Delete from Treeview widget
        table.delete(selected_item)

        messagebox.showinfo("Success", "Row deleted successfully.")
    else:
        messagebox.showerror("Error", "No row selected.")


def btn_add():

        if Temp_Input.get() == "" or Am_Input.get() == "" or pH_Input.get() == "" or Orp_Input.get() == "" or Wl_Input.get() == ""or BOD_Input.get() == ""or COD_Input.get() == ""or DO_Input.get() == "":
            messagebox.showerror("Error", "All field must be insert !!!")
        else:
            Temp = round(float(Temp_Input.get()), 2)
            pH = round(float(pH_Input.get()), 2)
            Am = round(float(Am_Input.get()), 2)
            Orp = round(float(Orp_Input.get()), 2)
            Wl = round(float(Wl_Input.get()), 2)
            BOD = round(float(BOD_Input.get()), 2)
            COD = round(float(COD_Input.get()), 2)
            DO = round(float(DO_Input.get()), 2)
            print(pH, Temp, Am, Orp, Wl,BOD,COD,DO)

            path = "D:\\Pycharm\\IML_Project\\new.xlsx"
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active
            row_values = [pH, Temp, Am, Orp, Wl,BOD,COD,DO]
            sheet.append(row_values)
            workbook.save("D:\\Pycharm\\IML_Project\\new.xlsx")
            load_data()

            # Clear inputs
            Temp_Input.delete(0, END)
            pH_Input.delete(0, END)
            Am_Input.delete(0, END)
            Orp_Input.delete(0, END)
            Wl_Input.delete(0, END)
            BOD_Input.delete(0, END)
            COD_Input.delete(0, END)
            DO_Input.delete(0, END)

            messagebox.showinfo("Success", "Data added successfully!")


def train_data():
    global trained_model

    data = pd.read_excel(
        "D:\\Pycharm\\IML_Project\\new.xlsx"
    )
    df = pd.DataFrame(data)
    print(df)
    check = df.isnull().sum()
    print(check)

    x = df.drop(columns=['pH'])

    print(x)

    y = df["pH"]

    print(y)

    x_train, x_test, y_train, y_test = train_test_split(x, y, test_size=0.3, random_state=0)

    lr = LinearRegression()
    lr.fit(x_train, y_train)

    trained_model = lr
    y_pred = lr.predict(x_test)

    r2 = r2_score(y_test, y_pred)
    rmse = np.sqrt(mean_squared_error(y_test, y_pred))

    print(r2)
    print(rmse)

    messagebox.showinfo("Training", "Model trained successfully!")

def predict():
    if trained_model is None:
        messagebox.showerror("Error", "Model is not trained. Please train the model first.")
        return

    if pH_Input.get() =="":
        # Get input values from the user interface (assuming inputs are text fields in your GUI)
        temp = float(Temp_Input.get())  # Assuming Temp_Input is the name of your input field
        ammonia = float(Am_Input.get())  # Ammonia input field
        orp = float(Orp_Input.get())  # ORP input field
        water_level = float(Wl_Input.get())  # Water level input field
        BOD = float(BOD_Input.get())
        COD = float(COD_Input.get())
        DO = float(DO_Input.get())
        # Create a DataFrame for the input values (as it is expected by the model)
        input_data = pd.DataFrame([[temp, ammonia, orp, water_level,BOD,COD,DO]],
                                  columns=['Temperature', 'Ammonia', 'ORP', 'Water_Level','BOD', 'COD', 'DO'])

        # Predict the pH value using the trained model
        prediction = trained_model.predict(input_data)

        # Display the prediction
        messagebox.showinfo("Prediction", f"Predicted pH value: {prediction[0]:.2f}")
    elif pH_Input.get() !="":
        messagebox.showinfo("Warning", "cannot do any prediction by entry ph value")


def plot_button():
    # Get the selected plot type from the radio buttons
    plot_type = plot_option.get()

    # Read the data into a DataFrame (assuming it's already loaded in 'df')
    data = pd.read_excel("D:\\Pycharm\\IML_Project\\new.xlsx")
    df = pd.DataFrame(data)
    # Check if we need to plot a Heatmap or Histogram
    if plot_type == 1:  # Histogram
        plt.figure(figsize=(10, 8))
        # Plot histogram for the selected feature (using X-Axis combo value)
        selected_feature = combo_X.get()  # Get selected feature for X-axis
        plt.hist(df[selected_feature], bins=20, color="skyblue", edgecolor="black", alpha=0.7)
        plt.title(f"Histogram of {selected_feature}", fontsize=12)
        plt.xlabel(col)
        plt.ylabel("Frequency")
        plt.show()

    elif plot_type == 2:  # Histogram
        plt.figure(figsize=(10, 8))
        # Plot boxplot for the selected feature (using X-Axis combo value)
        selected_feature = combo_X.get()  # Get selected feature for X-axis
        sb.boxplot(x=df[selected_feature], color="orange", fliersize=5, width=0.3)
        plt.title(f"Box Plot of {selected_feature}", fontsize=12)
        plt.show()

    elif plot_type == 3:  # Heatmap
        plt.figure(figsize=(10, 8))
        # Correlation matrix for plotting heatmap
        correlation_matrix = data.corr()
        sb.heatmap(correlation_matrix, annot=True, cmap="coolwarm", fmt=".2f")
        plt.title("Heatmap of Feature Correlations")
        plt.show()

    elif plot_type == 4:
        plt.figure(figsize=(10, 8))
        # Plot Scattor for the selected feature
        selected_feature = combo_Y.get()  # Get selected feature for y-axis
        sb.regplot(
            x=df["pH"], y=df[selected_feature],
            scatter_kws={"s": 100, "color": "green", "edgecolor": "black", "alpha": 0.75},
            line_kws={"color": "red", "linewidth": 2, "linestyle": "--"}
        )
        plt.title(f"pH vs of {selected_feature} with trend line", fontsize=12)
        plt.xlabel("pH")
        plt.ylabel(f"{selected_feature}")
        plt.show()
    else:
        messagebox.showinfo("No Plot Selected", "Please select either Heatmap or Histogram plot.")

# Create a Tkinter window
window = Tk()
window.geometry("1320x720")
window.title("Water Quality System")
window.resizable(width=False, height=False)

Main = Frame(window, background="#161C30")
Main.pack(fill="both", expand=True)
# Title
label = Label(
    Main,
    text="Water Quality System",
    font=("Arial", 32, "bold"),
    bg="#161C30",
    fg="white",
)
label.pack(pady=5)

content_frame = Frame(Main, background="#161C30")
content_frame.pack(fill="both", expand=True, padx=20, pady=8)

left_frame = Frame(content_frame, bg="#161C30")
left_frame.pack(side="left", fill="y", padx=10, pady=10, anchor="n")

# User Input Section (Left Side)
# User Input Section (Left Side)
frame_insert = LabelFrame(
    left_frame,
    text="User Insert",
    font=("Arial", 16, "bold"),
    padx=5,
    pady=5,
    bg="#161C30",
    fg="white",
)
frame_insert.pack(fill="x", padx=5, pady=5, anchor="n")

# Common font and entry width settings
entry_font = ("Arial", 10)  # Smaller font size
entry_width = 15  # Narrower width

# Temperature Input
Temp_label = Label(
    frame_insert,
    text="TEMPERATURE (°C):",
    font=("Arial", 10, "bold"),
    bg="#161C30",
    fg="white",
)
Temp_Input = Entry(frame_insert, font=entry_font, width=entry_width, borderwidth=3)
Temp_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
Temp_Input.grid(row=0, column=1, padx=5, pady=5, sticky="w")

# pH Value Input
pH_label = Label(
    frame_insert,
    text="pH VALUE:",
    font=("Arial", 10, "bold"),
    bg="#161C30",
    fg="white",
)
pH_Input = Entry(frame_insert, font=entry_font, width=entry_width, borderwidth=3)
pH_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")
pH_Input.grid(row=1, column=1, padx=5, pady=5, sticky="w")

# Ammonia Input
Am_label = Label(
    frame_insert,
    text="AMMONIA:",
    font=("Arial", 10, "bold"),
    bg="#161C30",
    fg="white",
)
Am_Input = Entry(frame_insert, font=entry_font, width=entry_width, borderwidth=3)
Am_label.grid(row=2, column=0, padx=5, pady=5, sticky="w")
Am_Input.grid(row=2, column=1, padx=5, pady=5, sticky="w")

# Orp Input
Orp_label = Label(
    frame_insert,
    text="ORP:",
    font=("Arial", 10, "bold"),
    bg="#161C30",
    fg="white",
)
Orp_Input = Entry(frame_insert, font=entry_font, width=entry_width, borderwidth=3)
Orp_label.grid(row=3, column=0, padx=5, pady=5, sticky="w")
Orp_Input.grid(row=3, column=1, padx=5, pady=5, sticky="w")

# Water Level Input
Wl_label = Label(
    frame_insert,
    text="WATER LEVEL:",
    font=("Arial", 10, "bold"),
    bg="#161C30",
    fg="white",
)
Wl_Input = Entry(frame_insert, font=entry_font, width=entry_width, borderwidth=3)
Wl_label.grid(row=0, column=2, padx=5, pady=5, sticky="w")
Wl_Input.grid(row=0, column=3, padx=5, pady=5, sticky="w")

# Conductivity Input
BOD_label = Label(
    frame_insert,
    text="BOD:",
    font=("Arial", 10, "bold"),
    bg="#161C30",
    fg="white",
)
BOD_Input = Entry(frame_insert, font=entry_font, width=entry_width, borderwidth=3)
BOD_label.grid(row=1, column=2, padx=5, pady=5, sticky="w")
BOD_Input.grid(row=1, column=3, padx=5, pady=5, sticky="w")

# Turbidity Input
COD_label = Label(
    frame_insert,
    text="COD:",
    font=("Arial", 10, "bold"),
    bg="#161C30",
    fg="white",
)
COD_Input = Entry(frame_insert, font=entry_font, width=entry_width, borderwidth=3)
COD_label.grid(row=2, column=2, padx=5, pady=5, sticky="w")
COD_Input.grid(row=2, column=3, padx=5, pady=5, sticky="w")

# Salinity Input
DO_label = Label(
    frame_insert,
    text="DO:",
    font=("Arial", 10, "bold"),
    bg="#161C30",
    fg="white",
)
DO_Input = Entry(frame_insert, font=entry_font, width=entry_width, borderwidth=3)
DO_label.grid(row=3, column=2, padx=5, pady=5, sticky="w")
DO_Input.grid(row=3, column=3, padx=5, pady=5, sticky="w")


# User Input Section (Left Side)
frame_button = LabelFrame(
    left_frame,
    text="Controls",
    font=("Arial", 16, "bold"),
    padx=5,
    pady=5,
    bg="#161C30",
    fg="white",
)
frame_button.pack(fill="x", padx=5, pady=5, anchor="n")

btn_add = Button(
    frame_button,
    text="ADD",
    font=("Arial", 12, "bold"),
    bg="Green",
    pady=5,
    width=10,
    fg="white",
    activebackground="lightgreen",
    command=btn_add
)
btn_add.grid(row=0, column=0, padx=5, sticky="ew")


btn_edit = Button(
    frame_button,
    text="EDIT",
    font=("Arial", 12, "bold"),
    bg="Blue",
    pady=5,
    width=10,
    fg="white",
    activebackground="lightblue",
    command=edit_button
)
btn_edit.grid(row=0, column=1, padx=5, sticky="ew")

btn_delete = Button(
    frame_button,
    text="DELETE",
    font=("Arial", 12, "bold"),
    bg="RED",
    pady=5,
    width=10,
    fg="white",
    activebackground="#FF7F7F",
    command=delete_row
)
btn_delete.grid(row=0, column=2, padx=5, sticky="ew")

btn_Predict = Button(
    frame_button,
    text="PREDICT",
    font=("Arial", 12, "bold"),
    bg="Gray",
    pady=5,
    width=10,
    fg="white",
    activebackground="lightgray",
    command=predict
)
btn_Predict.grid(row=0, column=3, padx=5, sticky="ew", columnspan=2)

btn_train = Button(
    frame_button,
    text="TRAIN",
    font=("Arial", 12, "bold"),
    bg="Yellow",
    pady=5,
    width=10,
    fg="Black",
    activebackground="#daee0f",
    command=train_data
)
btn_train.grid(row=1, column=0, padx=5, pady=5, sticky="ew", columnspan=4)

# Make Columns Responsive
frame_button.grid_columnconfigure(0, weight=1)
frame_button.grid_columnconfigure(1, weight=1)
frame_button.grid_columnconfigure(2, weight=1)
frame_button.grid_columnconfigure(3, weight=1)


# User Input Section (Left Side)
frame_plot = LabelFrame(
    left_frame,
    text="Plot Graph",
    font=("Arial", 16, "bold"),
    padx=20,
    pady=20,
    bg="#161C30",
    fg="white",
)
frame_plot.pack(fill="x", padx=10, pady=10, anchor="n")


plot_option = IntVar(value=1)

rd_histogram = Radiobutton(
    frame_plot,
    text="Histogram",
    font=("Arial", 12, "bold"),
    value=1,
    variable=plot_option,
    command=update_combobox_state,
    bg="#161C30",
    fg="white",
    activebackground="#1e243f",
    activeforeground="white",
    selectcolor="#161C30",
)
rd_histogram.grid(row=0, column=0, sticky="w", padx=10, pady=5)

rd_Box = Radiobutton(
    frame_plot,
    text="Box Plot",
    font=("Arial", 12, "bold"),
    value=2,
    command=update_combobox_state,
    variable=plot_option,
    bg="#161C30",
    fg="white",
    activebackground="#1e243f",
    activeforeground="white",
    selectcolor="#161C30",
)
rd_Box.grid(row=1, column=0, sticky="w", padx=10, pady=5)

rd_Heatmap = Radiobutton(
    frame_plot,
    text="Heatmap",
    font=("Arial", 12, "bold"),
    value=3,
    variable=plot_option,
    command=update_combobox_state,
    bg="#161C30",
    fg="white",
    activebackground="#1e243f",
    activeforeground="white",
    selectcolor="#161C30",
)
rd_Heatmap.grid(row=0, column=1, sticky="w", padx=10, pady=5)

rd_Scattor = Radiobutton(
    frame_plot,
    text="Scattor Plot",
    font=("Arial", 12, "bold"),
    value=4,
    variable=plot_option,
    command=update_combobox_state,
    bg="#161C30",
    fg="white",
    activebackground="#1e243f",
    activeforeground="white",
    selectcolor="#161C30",
)

rd_Scattor.grid(row=1, column=1, sticky="w", padx=10, pady=5)
# Labels for X-Axis and Y-Axis
label_x = Label(
    frame_plot, text="X-Axis", font=("Arial", 12, "bold"), bg="#161C30", fg="white"
)
label_x.grid(row=2, column=0, sticky="w", padx=10)

label_y = Label(
    frame_plot, text="Y-Axis", font=("Arial", 12, "bold"), bg="#161C30", fg="white"
)
label_y.grid(row=2, column=1, sticky="w", padx=10)

combo_X = Combobox(frame_plot, font=("Arial", 12), width=20, state="readonly")
combo_X["values"] = ("pH", "Temperature", "Ammonia", "ORP", "Water_Level","BOD","COD","DO")
combo_X.current(0)  # Set default value
combo_X.grid(row=3, column=0, sticky="w", padx=10, pady=10)

combo_Y = Combobox(frame_plot, font=("Arial", 12), width=20, state="readonly")
combo_Y["values"] = ("Temperature", "Ammonia", "ORP", "Water_Level","BOD","COD","DO")
combo_Y.current(0)  # Set default value
combo_Y.grid(row=3, column=1, sticky="w", padx=10, pady=10)

btn_Plot = Button(
    frame_plot,
    text="PLOT",
    font=("Arial", 12, "bold"),
    bg="#007BFF",
    fg="white",
    padx=10,
    pady=5,
    width=15,
    activebackground="#0056b3",
    command=plot_button
)
btn_Plot.grid(row=4, column=0, columnspan=2, pady=10)

frame_plot.grid_columnconfigure(0, weight=1)
frame_plot.grid_columnconfigure(1, weight=1)

combo_X.config(state="enabled")
combo_Y.config(state="disabled")

# Data Display Section
display_frame = LabelFrame(content_frame, text="Analyzed Data", font=("Arial", 16, "bold"), padx=20, pady=20, bg="#161C30", fg="white",bd=2,relief="solid")
display_frame.pack(fill="both", padx=10, pady=10, anchor="s", expand=True)

# Configure rows and columns in display_frame
display_frame.grid_rowconfigure(1, weight=1)
display_frame.grid_columnconfigure(0, weight=1)

# Search Frame
search_frame = Frame(display_frame, bg="#161C30")
search_frame.grid(row=0, column=0, sticky="ew", pady=5)

search_label = Label(search_frame, text="Search Option:", font=("Arial", 12, "bold"), bg="#161C30", fg="white")
search_label.grid(row=0, column=0, padx=5)

criteria_combo = Combobox(search_frame, values=["pH Value", "Temperature", 'Ammonia', 'ORP', "Water Level"], state="readonly", width=15)
criteria_combo.grid(row=0, column=1, padx=5)

btn_search = Button(search_frame, text="SEARCH", font=("Arial", 12, "bold"), bg="blue", fg="white")
btn_search.grid(row=0, column=3, padx=5)

# Treeview for displaying data
columns = ['pH', 'Temperature', 'Ammonia', 'ORP', 'Water Level','BOD', 'COD', 'DO']
table = Treeview(display_frame, columns=columns, show='headings')
# Add scrollbars
scroll_y = Scrollbar(display_frame, orient="vertical", command=table.yview)
scroll_x = Scrollbar(display_frame, orient="horizontal", command=table.xview)
table.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
table.bind("<ButtonRelease-1>", select_row)
# Position Treeview and scrollbars
table.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
scroll_y.grid(row=1, column=1, sticky="ns")
scroll_x.grid(row=2, column=0, sticky="ew")

# Set up Treeview headers
table.heading("pH", text="pH Value")
table.heading("Temperature", text="Temperature (°C)")
table.heading("Ammonia", text="Ammonia")
table.heading("ORP", text="ORP")
table.heading("Water Level", text="Water Level")
table.heading("BOD", text="BOD")
table.heading("COD", text="COD")
table.heading("DO", text="DO")


# Set column widths and alignment
for col in columns:
    table.column(col, anchor="center", width=150)


load_data()

window.mainloop()
